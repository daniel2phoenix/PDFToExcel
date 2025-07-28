import streamlit as st
import tabula

from datetime import datetime
from openpyxl import load_workbook
from PIL import Image


# variables
month2Number = {
  'Januar': [1, 'B'], 
  'Februar': [2, 'C'], 
  'März': [3, 'D'], 
  'April': [4, 'E'], 
  'Mai': [5, 'F'], 
  'Juni': [6, 'G'], 
  'Juli': [7, 'H'], 
  'August': [8, 'I'], 
  'September': [9, 'J'],
  'Oktober': [10, 'K'],
  'November': [11, 'L'],
  'Dezember': [12, 'M']
  }
monthList = ['Januar','Februar', 'März', 'April', 'Mai','Juni', 'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']

# function to build the sum of german numbers like 10.305,56€
def sumGermanNumbers(numbers):
  return sum(float(n.replace('.', '').replace(',', '.')) for n in numbers)


st.set_page_config(page_title='Excel Umwandler')
# Site Title
st.title('PDF in Excel einfügen')
# Define columns in the fiste row
col1, col2 = st.columns(2)

# load images form a pdf and excel file in the first row
with col1:
  pdf_image = Image.open('images/pdf.png')
  st.image(pdf_image, width=100, )

with col2:
  excel_image = Image.open('images/excel.png')
  st.image(excel_image, width=57)

col1_1, col2_2 = st.columns(2)

# show a file uploader in the second row column 1 and 2
with col1_1:
  try:
    uploaded_file = st.file_uploader('Wähle ein PDF Dokument aus:', type='pdf')
  except Exception as e:
    st.exception(f'Es werden nur PDFs zum Upload akzeptiert: {e}')

with col2_2:
  try:
    excel_file = st.file_uploader('Wähle die Excel-Datei aus:', type='xlsx')
  except Exception as e:
    st.exception(f'Es werden nur Excel-Dateien (.xlsx) zum Upload akzeptiert: {e}')


# input field for the year - default: the actual year
year = st.text_input('Geben sie das Jahr an:', value=datetime.now().year)
# selectionbox with 12 month to select
month = st.selectbox('Wähle den Monat aus, der in die Excel eingefügt werden soll:', monthList)
# Enter password Excel
password_excel = st.text_input('Geben Sie das Passwort für die Excel ein:')

# set file name and save excel file -
path_excel = st.text_input('Geben Sie den Speicherort an:')
new_path = path_excel + '/' + f'{year} Reporting Tool PxC Blomberg {month2Number[month][0]:02d}-{year}.xlsx'

if uploaded_file and month:
  st.markdown('---')
  # read tables from pdf file
  aramark_table = tabula.read_pdf(uploaded_file, pages='all', pandas_options={'header': None})
  try:
    row6 = aramark_table[0].iloc[0, month2Number[month][0]]
    row7 = sumGermanNumbers(aramark_table[0].iloc[[1, 3, 4, 5], month2Number[month][0]])
    row15 = sumGermanNumbers(aramark_table[0].iloc[[7, 8], month2Number[month][0]])
    row16 = aramark_table[0].iloc[6, month2Number[month][0]]
    row17 = aramark_table[0].iloc[2, month2Number[month][0]]
    row18 = aramark_table[0].iloc[12, month2Number[month][0]]
    row21 = aramark_table[0].iloc[13, month2Number[month][0]]
    row25 = aramark_table[0].iloc[15, month2Number[month][0]]
    row29 = sumGermanNumbers(aramark_table[0].iloc[[16, 18, 19, 20], month2Number[month][0]])
    row38 = aramark_table[0].iloc[17, month2Number[month][0]]
    row44 = aramark_table[1].iloc[4, month2Number[month][0]]
    row59 = aramark_table[2].iloc[5, month2Number[month][0]]
    row60 = aramark_table[2].iloc[3, month2Number[month][0]]
    row63 = aramark_table[3].iloc[10, month2Number[month][0]]
  except Exception as e:
    st.exception(f'Es gibt ein Problem mit der PDF, überprüfen sie ob das Format und die typische Anzahl der Tabellen passen: {e}')



if excel_file and st.button('In Excel speichern'):
  # load workbook from importet excel file
  wb = load_workbook(excel_file)
  ws = wb.active

  # delete sheet protection
  ws.protection.set_password(password_excel)
  ws.protection.sheet = False

  # add data form pdf to excel
  ws[f'{month2Number[month][1]}6'] = row6
  ws[f'{month2Number[month][1]}7'] = row7
  ws[f'{month2Number[month][1]}15'] = row15
  ws[f'{month2Number[month][1]}16'] = row16
  ws[f'{month2Number[month][1]}17'] = row17
  ws[f'{month2Number[month][1]}18'] = row18
  ws[f'{month2Number[month][1]}21'] = row21
  ws[f'{month2Number[month][1]}25'] = row25
  ws[f'{month2Number[month][1]}29'] = row29
  ws[f'{month2Number[month][1]}38'] = row38
  ws[f'{month2Number[month][1]}44'] = row44
  ws[f'{month2Number[month][1]}59'] = row59
  ws[f'{month2Number[month][1]}60'] = row60
  ws[f'{month2Number[month][1]}63'] = row63

  # set sheet protection
  ws.protection.set_password(password_excel)
  ws.protection.sheet = True
    
  try:
    wb.save(new_path)
    print(f'Excel gespeichtert:,{new_path}')
  except Exception as e:
    print(f'Ein Fehler beim speichern ist aufgetreten. Evtl ist der Dateiname schon vorhanden. Das Überschreiben ist nicht möglich. Bitte löschen Sie die Datei oder ändern den Namen: {e}')



